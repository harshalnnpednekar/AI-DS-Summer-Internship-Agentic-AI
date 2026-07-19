import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _color_cell(cell, pct):
    """Apply green/yellow/red color based on attendance percentage."""
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if pct >= 75:
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell.font = Font(color="006100", bold=True)
    elif pct >= 60:
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        cell.font = Font(color="9C5700", bold=True)
    else:
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.font = Font(color="9C0006", bold=True)

THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

class ExcelSyncAgent:
    def generate_subject_sheet(self, all_lectures_data, name_lookup, faculty_name, class_name, subject_name, session_type, total_students):
        """
        Generate the subject-wise attendance matrix in-memory.
        
        Parameters:
        - all_lectures_data: list of dicts with lecture_date, time_slot, absentee_roll_numbers
        - name_lookup: dict of {roll_number_str -> student_name}
        - faculty_name: str
        - class_name: str
        - subject_name: str
        - session_type: 'Theory' or 'Practical'
        - total_students: int (used to generate roll numbers 1..N)
        
        Returns: bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Log"  # type: ignore

        all_rolls = [str(i) for i in range(1, total_students + 1)]
        num_lectures = len(all_lectures_data)
        
        # Columns: A (Roll), B (Name), C.. (Lectures), Last (Percentage)
        total_cols = max(3, 2 + num_lectures + 1)
        pct_col_idx = 2 + num_lectures + 1
        pct_col_letter = get_column_letter(pct_col_idx)

        # Row 1 - Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)  # type: ignore
        ws['A1'] = f"Class: {class_name}  |  Faculty: {faculty_name}  |  Subject: {subject_name} ({session_type})"  # type: ignore
        ws['A1'].font = Font(bold=True, size=13)  # type: ignore
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")  # type: ignore
        ws.row_dimensions[1].height = 25  # type: ignore

        # Row 2 - Headers
        ws['A2'] = "Roll No"  # type: ignore
        ws['A2'].font = Font(bold=True)  # type: ignore
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")  # type: ignore
        ws['B2'] = "Name"  # type: ignore
        ws['B2'].font = Font(bold=True)  # type: ignore
        ws['B2'].alignment = Alignment(horizontal="center", vertical="center")  # type: ignore
        
        # Lecture Headers
        for i, lec in enumerate(all_lectures_data):
            col = i + 3
            col_letter = get_column_letter(col)
            
            lecture_date = lec['lecture_date']
            date_str = lecture_date.strftime("%Y-%m-%d") if hasattr(lecture_date, 'strftime') else str(lecture_date)
            try:
                day_str = lecture_date.strftime("%A")
            except Exception:
                day_str = ""
            time_slot = lec.get('time_slot', '') or ''
            header_text = f"{date_str}\n{day_str}\n{time_slot}"
            
            ws[f'{col_letter}2'] = header_text  # type: ignore
            ws[f'{col_letter}2'].font = Font(bold=True)  # type: ignore
            ws[f'{col_letter}2'].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")  # type: ignore
            ws[f'{col_letter}2'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # type: ignore
            ws.column_dimensions[col_letter].width = 14  # type: ignore

        # Percentage Header
        ws[f'{pct_col_letter}2'] = "Attendance %"  # type: ignore
        ws[f'{pct_col_letter}2'].font = Font(bold=True)  # type: ignore
        ws[f'{pct_col_letter}2'].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")  # type: ignore
        ws[f'{pct_col_letter}2'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # type: ignore
        ws.row_dimensions[2].height = 40  # type: ignore

        # Row data
        for idx, roll in enumerate(all_rolls):
            r = idx + 3
            ws[f'A{r}'] = roll  # type: ignore
            ws[f'A{r}'].alignment = Alignment(horizontal="center", vertical="center")  # type: ignore
            ws[f'B{r}'] = name_lookup.get(roll, "")  # type: ignore
            
            total_p = 0
            total_a = 0
            
            for i, lec in enumerate(all_lectures_data):
                col = i + 3
                col_letter = get_column_letter(col)
                absentees = set(str(abs_r).strip() for abs_r in lec.get('absentee_roll_numbers', []))
                
                mark = 'A' if roll in absentees else 'P'
                cell = ws[f'{col_letter}{r}']  # type: ignore
                cell.value = mark
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if mark == 'P':
                    cell.font = Font(color="006100", bold=True)
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    total_p += 1
                else:
                    cell.font = Font(color="9C0006", bold=True)
                    cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
                    total_a += 1
                    
            # Percentage
            total_lec = total_p + total_a
            pct_cell = ws[f'{pct_col_letter}{r}']  # type: ignore
            if total_lec > 0:
                pct = round((total_p / total_lec) * 100)
                pct_cell.value = f"{pct}%"
                _color_cell(pct_cell, pct)
            else:
                pct_cell.value = "-"
                pct_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions['A'].width = 10  # type: ignore
        ws.column_dimensions['B'].width = 28  # type: ignore
        ws.column_dimensions[pct_col_letter].width = 14  # type: ignore

        for r in range(1, ws.max_row + 1):  # type: ignore
            for c in range(1, ws.max_column + 1):  # type: ignore
                ws.cell(row=r, column=c).border = THIN  # type: ignore

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def generate_master_sheet(self, class_name, students_list, class_data):
        """
        Generates master sheet with all subjects for a class in-memory.
        Returns: bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Master Report"  # type: ignore

        subject_keys = list(class_data.keys())

        # Row 1 – Title (spans all cols)
        total_cols = 2 + len(subject_keys) * 2 + len(subject_keys) + 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(5, total_cols))  # type: ignore
        ws['A1'] = f"{class_name}  —  MASTER ATTENDANCE REPORT"  # type: ignore
        ws['A1'].font = Font(bold=True, size=14)  # type: ignore
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")  # type: ignore
        ws.row_dimensions[1].height = 28  # type: ignore

        # Row 2 – headers
        ws['A2'] = "Roll No"  # type: ignore
        ws['B2'] = "Name"  # type: ignore
        for cell in [ws['A2'], ws['B2']]:  # type: ignore
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        col_idx = 3
        subject_cols = {}

        for key in subject_keys:
            subject = class_data[key]["subject_name"]

            t_col = get_column_letter(col_idx)
            ws[f'{t_col}2'] = f"{subject}\nTheory\n({class_data[key]['total_theory']})"  # type: ignore
            ws[f'{t_col}2'].font = Font(bold=True)  # type: ignore
            ws[f'{t_col}2'].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")  # type: ignore
            subject_cols[key] = {"theory_col": t_col}
            col_idx += 1

            p_col = get_column_letter(col_idx)
            ws[f'{p_col}2'] = f"{subject}\nPractical\n({class_data[key]['total_practical']})"  # type: ignore
            ws[f'{p_col}2'].font = Font(bold=True)  # type: ignore
            ws[f'{p_col}2'].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")  # type: ignore
            subject_cols[key]["prac_col"] = p_col
            col_idx += 1

        # Percentage summary block header
        pct_start = col_idx
        pct_end = col_idx + len(subject_keys)  # subject pcts + overall
        
        ws.cell(row=1, column=pct_start).value = "Attendance Percentage  :  Summary"  # type: ignore
        ws.cell(row=1, column=pct_start).font = Font(bold=True, size=11)  # type: ignore
        ws.cell(row=1, column=pct_start).alignment = Alignment(horizontal="center", vertical="center")  # type: ignore
        ws.merge_cells(start_row=1, start_column=pct_start, end_row=1, end_column=pct_end)  # type: ignore

        for key in subject_keys:
            subject = class_data[key]["subject_name"]
            s_col = get_column_letter(col_idx)
            ws[f'{s_col}2'] = f"{subject}\n%"  # type: ignore
            ws[f'{s_col}2'].font = Font(bold=True)  # type: ignore
            ws[f'{s_col}2'].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")  # type: ignore
            ws[f'{s_col}2'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # type: ignore
            subject_cols[key]["pct_col"] = s_col
            col_idx += 1

        o_col = get_column_letter(col_idx)
        ws[f'{o_col}2'] = "Overall\n%"  # type: ignore
        ws[f'{o_col}2'].font = Font(bold=True)  # type: ignore
        ws[f'{o_col}2'].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")  # type: ignore
        ws[f'{o_col}2'].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # type: ignore
        ws.row_dimensions[2].height = 50  # type: ignore

        # Student rows
        row = 3
        for student in sorted(students_list, key=lambda x: str(x['roll_number'])):
            roll = str(student['roll_number'])
            ws[f'A{row}'] = roll  # type: ignore
            ws[f'B{row}'] = student.get('name', '')  # type: ignore
            ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")  # type: ignore

            total_attended_all = 0
            total_sessions_all = 0

            for key in subject_keys:
                t_total = class_data[key]["total_theory"]
                p_total = class_data[key]["total_practical"]

                absent_t = class_data[key]["absentees"].get(roll, {}).get("Theory", 0)
                absent_p = class_data[key]["absentees"].get(roll, {}).get("Practical", 0)

                attended_t = t_total - absent_t
                attended_p = p_total - absent_p

                t_col = subject_cols[key]["theory_col"]
                p_col = subject_cols[key]["prac_col"]
                pct_col = subject_cols[key]["pct_col"]

                ws[f'{t_col}{row}'] = f"{attended_t}/{t_total}" if t_total > 0 else "-"  # type: ignore
                ws[f'{p_col}{row}'] = f"{attended_p}/{p_total}" if p_total > 0 else "-"  # type: ignore
                for c in [t_col, p_col]:
                    ws[f'{c}{row}'].alignment = Alignment(horizontal="center", vertical="center")  # type: ignore

                subj_total = t_total + p_total
                subj_attended = attended_t + attended_p
                total_attended_all += subj_attended
                total_sessions_all += subj_total

                if subj_total > 0:
                    pct = round((subj_attended / subj_total) * 100)
                    pct_cell = ws[f'{pct_col}{row}']  # type: ignore
                    pct_cell.value = f"{pct}%"
                    _color_cell(pct_cell, pct)
                else:
                    ws[f'{pct_col}{row}'] = "-"  # type: ignore

            if total_sessions_all > 0:
                overall_pct = round((total_attended_all / total_sessions_all) * 100)
                o_cell = ws[f'{o_col}{row}']  # type: ignore
                o_cell.value = f"{overall_pct}%"
                _color_cell(o_cell, overall_pct)
            else:
                ws[f'{o_col}{row}'] = "-"  # type: ignore

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 10  # type: ignore
        ws.column_dimensions['B'].width = 28  # type: ignore
        for i in range(3, col_idx + 1):
            ws.column_dimensions[get_column_letter(i)].width = 13  # type: ignore

        # Borders
        for r in range(1, ws.max_row + 1):  # type: ignore
            for c in range(1, ws.max_column + 1):  # type: ignore
                ws.cell(row=r, column=c).border = THIN  # type: ignore

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

excel_sync_agent = ExcelSyncAgent()
